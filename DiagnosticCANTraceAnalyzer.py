import re
import pandas as pd
import configparser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def parse_log(log_lines, request_id, response_id, project_type):
    parsed_data = []
    requests = []  # Store request lines
    responses = []  # Store response lines
    current_data = None
    current_time = None
    current_can_id = None
    delta_time = None
    total_length = 0  # Initialize total_length for multi-line data
    is_processing_multiline = False  # Track if multi-line data is being processed

    for line in log_lines:
        # Ignore lines containing specific keywords or phrases
        if (
            "Status:chip status error active" in line or
            "Statistic: D" in line or
	        "Statistic: D" in line or
            "Begin TriggerBlock" in line or
	        "internal events logged" in line or
            "// version" in line or
            "base hex" in line or
            "timestamps absolute" in line or
            "date" in line or
            "Rx   d 3 30 00 00" in line or
            "Tx   d 3 30 00 00" in line or
            "Rx   d 8 30 00 00" in line or
            "Tx   d 8 30 00 00" in line or
            "Rx   d 8 30 FF 0A" in line or
            "Rx   d 8 30 08 0A" in line or
            "Tx   d 8 30 FF 0A" in line or
	        "1 0 8  8 30 00 14" in line 
        ):
            continue  # Skip this line
        # Initialize match to None
        match = None
        parsed_successfully = False
        if project_type == "CAN":
            # Match the CAN log line format
            match = re.match(r"\s*(\d+\.\d+)\s+\d+\s+(\S+)\s+(Rx|Tx)\s+d\s+\d+\s+((?:[0-9A-F]{2}\s?)+)", line)
            if match:
                time = float(match.group(1))
                can_id = match.group(2)
                direction = match.group(3)  # Rx or Tx
                data = match.group(4).strip().split()
                parsed_successfully = True  # Set the flag to True if parsing is successful
        elif project_type == "CANFD":
            # Split the line into parts
            parts = line.split()
            try:
                # Extract fields based on their positions
                time = float(parts[0])  # Timestamp
                protocol = parts[1]     # Protocol (CANFD)
                can_id = parts[4]       # CAN ID
                data_length = parts[5:8]  # Data length fields (e.g., "1 0 8")
                framelength = int(parts[8])
                data = parts[8:8+framelength+1]  # Data bytes (e.g., "8 02 7e 00 cc cc cc cc cc")
                additional_data = parts[9+framelength:]  # Remaining data (e.g., "102031 130 323000 ...")
                parsed_successfully = True  # Set the flag to True if parsing is successful
            except (IndexError, ValueError):
                # Handle cases where parsing fails (e.g., malformed data)
                parsed_successfully = False

            # Define a regex pattern to match the CANFD log line format
            #match=re.match(r"\s*(\d+\.\d+)\s+CANFD\s+\d+\s+Rx\s+(\d+)\s+(\S+)\s+\d+\s+\d+\s+\d+\s+((?:[0-9A-F]{2}\s?)+)\s", line)
            #print(f"time {time} ")
            #print(f"data_length {data_length} ")
            #print(f"framelength {framelength} ")
            #print(f"data {data} ")
        if parsed_successfully:
            # Calculate delta time
            if current_time is not None:
                delta_time = round(time - current_time, 6)
            else:
                delta_time = 0.0
            # Handle multi-line data continuation for both requests and responses
            if (project_type == "CAN" and data[0] == "10") or (project_type == "CANFD" and int(data[0]) > 8 and not is_processing_multiline):
                # If there is an incomplete multi-line message, append it to parsed_data
                if is_processing_multiline and current_data and len(current_data) < total_length:
                    parsed_data.append([current_time, delta_time, current_can_id, " ".join(current_data) + " (Incomplete Data)"])
                
                if project_type == "CAN":
                    total_length = int(data[1], 16)
                    current_data = data[2:]
                    is_processing_multiline = True
                elif project_type == "CANFD" and  (int(data[0]) == 64) and ((int(f"{int(data[1], 16) & 0x0F:01X}{data[2]}", 16)) >= 62):
                    # Combine bytes for total length in CANFD
                    total_length  = int(f"{int(data[1], 16) & 0x0F:01X}{data[2]}", 16) # Combine bytes for length in CANFD
                    # Include all bytes from the current frame up to the total length
                    #print(f"data is {data[3:]} ")
                    current_data = data[3:]
                    is_processing_multiline = True
                elif project_type == "CANFD" and  (int(data[0]) <= 64) and ((int(f"{int(data[1], 16) & 0x0F:01X}{data[2]}", 16)) < 62):
                    # Combine bytes for total length in CANFD
                    total_length  = int(f"{int(data[1], 16) & 0x0F:01X}{data[2]}", 16)  # Combine bytes for length in CANFD
                    # Include all bytes from the current frame up to the total length
                    current_data = data[3:3+total_length]
                    parsed_data.append([time, delta_time, can_id, " ".join(current_data)])
                    current_data = None
                    #print(f"data was {current_data} ")
                    is_processing_multiline = False
                current_time = time
                current_can_id = can_id
            elif ((project_type == "CAN") and data[0] in ["21", "22", "23", "24", "25", "26", "27", "28", "29", "2a", "2b", "2c", "2e", "2f"]) or((project_type == "CANFD") and data[1] in ["21", "22", "23", "24", "25", "26", "27", "28", "29", "2a", "2b", "2c", "2e", "2f"]):  # Continuation of multi-line data
                if is_processing_multiline and current_data is not None:
                    if (project_type == "CAN"):
                        current_data.extend(data[1:])
                    elif (project_type == "CANFD"):
                        current_data.extend(data[2:])
                        #print(f"continued data was {current_data} ")
            else:  # Single-line data
                # If there is an incomplete multi-line message, append it to parsed_data
                if is_processing_multiline and current_data and len(current_data) < total_length:
                    parsed_data.append([current_time, delta_time, current_can_id, " ".join(current_data) + " (Incomplete Data)"])
                    current_data = None
                    is_processing_multiline = False
                # Handle multi-line data continuation for both requests and responses
                if project_type == "CAN":
                    data_length = int(data[0], 16)
                    real_data = data[1:data_length + 1]
                elif (project_type == "CANFD" and int(data[0]) <= 8):
                    data_length = int(data[1] , 16)   # First byte indicates the actual data length
                    real_data = data[2:data_length + 1+1]
                    #print(f"data {real_data} ")

                parsed_data.append([time, delta_time, can_id, " ".join(real_data)])
                current_data = None
                is_processing_multiline = False

            # If multi-line data is complete
            if is_processing_multiline and current_data and len(current_data) >= total_length:
                if project_type == "CANFD":
                    # Extract the actual data length from the first two bytes of the frame
                    #actual_data_length = int(f"{int(current_data[0], 16) & 0x0F:01X}{current_data[1]}", 16)
                    #real_data = current_data[2:2 + actual_data_length]
                    real_data = current_data[:total_length]
                else:
                    real_data = current_data[:total_length]

                parsed_data.append([current_time, delta_time, current_can_id, " ".join(real_data)])
                current_data = None
                is_processing_multiline = False

            current_time = time

            # Check if it's a request or response
            if can_id == request_id:  # Request
                requests.append((time, can_id, data))
            elif can_id == response_id:  # Response
                responses.append((time, can_id, data))

    # Check for incomplete multi-line data at the end
    if is_processing_multiline and current_data and len(current_data) < total_length:
        parsed_data.append([current_time, delta_time, current_can_id, " ".join(current_data) + " (Incomplete Data)"])

    # Check for missing responses
    for i in range(len(parsed_data) - 1):
        if parsed_data[i][2] == request_id and parsed_data[i + 1][2] != response_id:
            parsed_data[i][3] += " (Missing Response)"

    return parsed_data


def export_to_excel(parsed_data, output_file, ddl_file, enable_failure_identification):
    df = pd.DataFrame(parsed_data, columns=["Time", "DeltaTime", "CAN_ID", "Data"])
    if enable_failure_identification:
        # Load the CSV file and filter the sheet starting with "DiagnosisDataList"
        # Check if the file is Excel or CSV
        if ddl_file.endswith(".xlsx") or ddl_file.endswith(".xls"):
            # Load the Excel file and get the sheet names
            excel_file = pd.ExcelFile(ddl_file)
            # Find the sheet name that starts with "DiagnosisDataList"
            sheet_name = next((name for name in excel_file.sheet_names if name.startswith("DiagnosisDataList")), None)
            #print(sheet_name)
            if sheet_name is None:
                raise ValueError("No sheet starting with 'DiagnosisDataList' found in the Excel file.")
            # Load the data from the selected sheet
            diagnosis_data = pd.read_excel(ddl_file, sheet_name=sheet_name)
        else:
            # Load the CSV file with semicolon delimiter, skipping rows before the column names
            diagnosis_data = pd.read_csv(ddl_file, delimiter=";", header=1)  # Adjust rows to skip
        #print(diagnosis_data.columns)
        dtc_mapping = diagnosis_data[["DTC_VALUE", "FW_NAME"]].copy()
        # Convert DTC_VALUE to match the format in the Data column
        dtc_mapping["DTC_VALUE"] = dtc_mapping["DTC_VALUE"].str.replace("0x", "").str.upper()
        # Convert parsed data to a DataFrame
        df = pd.DataFrame(parsed_data, columns=["Time", "DeltaTime", "CAN_ID", "Data"])

        def parse_dtc(data):
            if data.startswith("59 02"):
                status_bit_mapping = {  # Map status bits to their descriptions
                    0x01: "Test Failed",
                    0x02: "Test Failed This Operation Cycle",
                    0x04: "Pending DTC",
                    0x08: "Confirmed DTC",
                    0x10: "Test Not Completed Since Last Clear",
                    0x20: "Test Failed Since Last Clear",
                    0x40: "Test Not Completed This Operation Cycle",
                    0x80: "Warning Indicator Requested",
                }
                status_to_dtc = {bit: [] for bit in status_bit_mapping}  # Initialize a dictionary to group DTCs by status bits
                bytes_data = data.split()
                i = 3  # Start from the 3rd byte (index 2)
                while i + 3 <= len(bytes_data):  # Ensure there are enough bytes for DTC and status
                    # Normalize the DTC number from Data to uppercase
                    dtc_number = "".join(bytes_data[i:i+3]).upper()  # 3-byte DTC number (e.g., "C10004")

                    # Match the normalized DTC number with the normalized DTC_VALUE in the mapping
                    failure_name = dtc_mapping.loc[dtc_mapping["DTC_VALUE"] == dtc_number, "FW_NAME"].values
                    failure_detail = f"{failure_name[0]}" if len(failure_name) > 0 else "Unknown Failure"

                    # Extract the status byte and decode it
                    status_byte = int(bytes_data[i+3], 16)  # Status byte
                    for bit, description in status_bit_mapping.items():
                        if status_byte & bit:  # Check if the bit is set
                            status_to_dtc[bit].append(f"DTC: {dtc_number}, Failure: {failure_detail}")
                    i += 4  # Move to the next DTC (3 bytes for DTC + 1 byte for status)

                # Build the remark by grouping DTCs under each status bit
                remarks = []
                for bit, description in status_bit_mapping.items():
                    if status_to_dtc[bit]:  # If there are DTCs for this status bit
                        remarks.append(f"{description}:\n" + "\n".join(status_to_dtc[bit]))
                return "\n\n".join(remarks)  # Separate each status group with a blank line
            return ""
        df["Remark"] = df["Data"].apply(parse_dtc)

    # Export to Excel
    df.to_excel(output_file, index=False)
    print(f"Data exported to {output_file}")

    # Highlight cells with issues
    highlight_issues(output_file)


def highlight_issues(output_file):
    # Load the Excel file
    wb = load_workbook(output_file)
    ws = wb.active

    # Define a red fill for highlighting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Iterate through the "Data" column to find issues
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if "Missing Response" in str(cell.value) or "Incomplete Data" in str(cell.value):
                cell.fill = red_fill

    # Save the updated Excel file
    wb.save(output_file)
    #print(f"Highlighted issues in {output_file}")


def main():
    # Read configuration from config.properties
    config = configparser.ConfigParser()
    config.read("config.properties")

    input_file = config.get("FILES", "input_file")
    output_file = config.get("FILES", "output_file")
    request_id = config.get("CAN_IDS", "request_id")
    response_id = config.get("CAN_IDS", "response_id")
    project_type = config.get("PROJECT", "type")  # CAN or CANFD
    ddl_file = config.get("DDL", "DiagnosisDataListPath")
    enable_failure_identification = config.getboolean("SETTINGS", "enable_failure_identification", fallback=False)

    # Read the input file
    with open(input_file, "r") as file:
        log_lines = file.readlines()

    # Parse the log lines
    parsed = parse_log(log_lines, request_id, response_id, project_type)
    #(f"Parsed log file:\n{parsed}")

    # Export the parsed data to an Excel file
    export_to_excel(parsed, output_file, ddl_file, enable_failure_identification)


if __name__ == "__main__":
    main()